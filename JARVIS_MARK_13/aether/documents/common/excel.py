"""
documents/common/excel.py

Handles Excel (.xlsx) workbook creation, reading, and cell writes.
Platform-independent business logic using openpyxl.
"""

from pathlib import Path
from typing import Any, List
import openpyxl
from openpyxl import Workbook

from aether.documents.common.exceptions import SheetNotFoundError, InvalidCellFormatError, DocumentCorruptedError

def create_excel_workbook(
    file_path: Path,
    sheet_name: str | None = None,
    overwrite: bool = False
) -> str:
    """
    Creates a new Excel (.xlsx) workbook with a specified active worksheet name.
    """
    if file_path.exists() and not overwrite:
        raise FileExistsError(f"File already exists: {file_path}")
        
    # Create parent directories if missing
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name.strip() if sheet_name and sheet_name.strip() else "Sheet1"
        wb.save(str(file_path))
        return str(file_path.resolve())
    except Exception as e:
        raise DocumentCorruptedError(f"Failed to create Excel workbook: {e}")

def read_excel_workbook(
    file_path: Path,
    sheet_name: str | None = None,
    cell_range: str | None = None
) -> List[List[Any]]:
    """
    Reads cells from a worksheet. If cell_range is None, reads the entire sheet.
    Returns cells as a list of lists of values.
    """
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    if not file_path.is_file():
        raise IsADirectoryError(f"Path is a directory: {file_path}")
        
    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
    except Exception as e:
        raise DocumentCorruptedError(f"Failed to load Excel workbook: {e}")
        
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Worksheet '{sheet_name}' not found.")
        ws = wb[sheet_name]
    else:
        ws = wb.active
        
    data = []
    try:
        if cell_range:
            # Evaluate using openpyxl range slicer
            cells = ws[cell_range]
            # If cells is tuple of tuples (representing range grid)
            if isinstance(cells, tuple):
                if len(cells) > 0 and isinstance(cells[0], tuple):
                    for row in cells:
                        row_vals = [cell.value for cell in row]
                        data.append(row_vals)
                else:
                    # Single row/column tuple
                    row_vals = [cell.value for cell in cells]
                    data.append(row_vals)
            else:
                # Single Cell object
                data.append([cells.value])
        else:
            # Default: Read entire grid
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
        return data
    except Exception as e:
        # Check standard coordinate issues
        if "coordinate" in str(e).lower() or "range" in str(e).lower():
            raise InvalidCellFormatError(f"Invalid range or cell coordinate format '{cell_range}': {e}")
        raise DocumentCorruptedError(f"Failed to read Excel workbook: {e}")

def write_excel_cell(
    file_path: Path,
    sheet_name: str | None = None,
    cell: str = "A1",
    value: Any = None
) -> str:
    """
    Writes a value to a cell in a worksheet. Creates the sheet if missing.
    """
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
        
    try:
        wb = openpyxl.load_workbook(str(file_path))
    except Exception as e:
        raise DocumentCorruptedError(f"Failed to load Excel workbook: {e}")
        
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
    else:
        ws = wb.active
        
    try:
        ws[cell] = value
    except Exception as e:
        raise InvalidCellFormatError(f"Invalid cell coordinate format '{cell}': {e}")
        
    try:
        wb.save(str(file_path))
        return f"Successfully updated cell '{cell}' with value '{value}' in worksheet '{ws.title}'."
    except Exception as e:
        raise DocumentCorruptedError(f"Failed to save Excel workbook: {e}")
